from openpyxl import Workbook
from openpyxl.styles import Font


wb = Workbook()
wb.remove(wb.active) 

language_sheet = wb.create_sheet(title="Language")
capital_sheet = wb.create_sheet(title="Capital")

lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code = ['KA', 'TS', 'TN']

def set_headers(sheet, headers):
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

set_headers(language_sheet, ["State", "Language", "Code"])
for i in range(2, 5):
    language_sheet.append([state[i - 2], lang[i - 2], code[i - 2]])

set_headers(capital_sheet, ["State", "Capital", "Code"])
for i in range(2, 5):
    capital_sheet.append([state[i - 2], capital[i - 2], code[i - 2]])

wb.save("demo.xlsx")

search_code = input("Enter state code for finding capital or language: ")
for sheet in wb:
    if sheet.title.lower() in ["language", "capital"]:
        for row in sheet.iter_rows(min_row=2, max_row=4, min_col=3, max_col=3):
            for cell in row:
                if cell.value == search_code:
                    print(
                        f"Corresponding {sheet.title.lower()} for code {search_code} is {row[1].value}")
wb.close()
